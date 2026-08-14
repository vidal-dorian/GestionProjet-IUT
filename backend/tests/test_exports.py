import io

from openpyxl import load_workbook

from tests.helpers import add_approved_member


def setup_authenticated_account(client, email="alice@test.local"):
    client.headers["X-Dev-Email"] = email
    project = client.post("/api/projects", json={"name": "Projet Export"}).json()
    return project, email


def test_export_my_entries_requires_authentication(client):
    project = client.post("/api/projects", json={"name": "Projet Export Sans Session"}).json()
    response = client.get(f"/api/projects/{project['id']}/time-entries/export")
    assert response.status_code == 401


def test_export_my_entries_without_session_for_project_returns_401(client):
    response = client.get("/api/projects/999/time-entries/export")
    assert response.status_code == 401


def test_export_my_entries_returns_xlsx_with_entries_sheet(client):
    project, _ = setup_authenticated_account(client)
    client.post(
        f"/api/projects/{project['id']}/time-entries",
        json={"date": "2026-08-13", "duration_hours": 2, "description": "Dev"},
    )

    response = client.get(f"/api/projects/{project['id']}/time-entries/export")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert "Projet_Export" in response.headers["content-disposition"]
    assert "alice_test_local" in response.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(response.content))
    assert "Entrées" in wb.sheetnames
    ws = wb["Entrées"]
    assert ws["A1"].value == "Date"
    assert ws.max_row == 2  # header + 1 entry


def test_export_my_entries_with_no_entries_has_no_chart_sheets(client):
    project, _ = setup_authenticated_account(client)

    response = client.get(f"/api/projects/{project['id']}/time-entries/export")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Entrées"]


def test_export_project_for_unknown_project_returns_404(client):
    client.headers["X-Dev-Email"] = "alice@test.local"
    response = client.get("/api/projects/999/export")
    assert response.status_code == 404


def test_export_project_returns_xlsx_with_summary_and_entries(client):
    project, _ = setup_authenticated_account(client)

    client.post(
        f"/api/projects/{project['id']}/time-entries",
        json={"date": "2026-08-13", "duration_hours": 2, "description": "Alice"},
    )
    add_approved_member(project["id"], "bob@test.local")
    client.headers["X-Dev-Email"] = "bob@test.local"
    client.post(
        f"/api/projects/{project['id']}/time-entries",
        json={"date": "2026-08-14", "duration_hours": 3, "description": "Bob"},
    )

    response = client.get(f"/api/projects/{project['id']}/export")
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(response.content))
    assert "Synthèse" in wb.sheetnames
    assert "Entrées" in wb.sheetnames
    assert "Heures par membre" in wb.sheetnames
    assert "Répartition par sprint" in wb.sheetnames
    assert "Évolution temporelle" in wb.sheetnames

    summary = wb["Synthèse"]
    summary_rows = {row[0].value: row[1].value for row in summary.iter_rows(min_row=2)}
    assert summary_rows["Total heures"] == 5.0
    assert summary_rows["Nombre de contributeurs"] == 2
    assert summary_rows["Nombre d'entrées"] == 2

    entries_sheet = wb["Entrées"]
    assert entries_sheet["A1"].value == "Date"
    assert entries_sheet["B1"].value == "Compte"
    assert entries_sheet.max_row == 3  # header + 2 entries


def test_export_project_with_no_entries_has_no_chart_sheets(client):
    client.headers["X-Dev-Email"] = "alice@test.local"
    project = client.post("/api/projects", json={"name": "Projet Export Vide"}).json()

    response = client.get(f"/api/projects/{project['id']}/export")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    assert set(wb.sheetnames) == {"Synthèse", "Entrées"}
