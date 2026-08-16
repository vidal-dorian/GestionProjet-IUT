import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app import models

ENTRY_HEADERS = ["Date", "Durée (h)", "Description", "Sprint", "Issue", "Catégorie"]

# Neutralise l'injection de formule (CSV/Excel injection) : un tableur interprète comme
# formule toute cellule texte commençant par l'un de ces caractères. Ces valeurs (description
# de saisie, noms de sprint/catégorie/projet...) sont fournies par n'importe quel membre du
# projet, donc non fiables — on les préfixe d'une apostrophe pour forcer un affichage littéral.
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _safe_row(row: list) -> list:
    return [_safe_cell(value) for value in row]


def _entry_row(entry: models.TimeEntry) -> list:
    return _safe_row(
        [
            entry.date,
            entry.duration_hours,
            entry.description,
            entry.sprint.name if entry.sprint else "",
            f"#{entry.github_issue.number} {entry.github_issue.title}" if entry.github_issue else "",
            entry.category.name if entry.category else "",
        ]
    )


def _write_entries_sheet(ws: Worksheet, entries: list[models.TimeEntry], *, include_account: bool) -> None:
    headers = list(ENTRY_HEADERS)
    if include_account:
        headers.insert(1, "Compte")
    ws.append(headers)

    for entry in entries:
        row = _entry_row(entry)
        if include_account:
            row.insert(1, _safe_cell(entry.account_email))
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)


def _write_summary_table(
    ws: Worksheet, start_col: int, start_row: int, data: list[tuple[str, float]], *, label_header: str = "Libellé"
) -> None:
    """Écrit un petit tableau libellé/heures à partir de (start_row, start_col)."""
    ws.cell(row=start_row, column=start_col, value=label_header)
    ws.cell(row=start_row, column=start_col + 1, value="Heures")
    for offset, (label, hours) in enumerate(data, start=1):
        ws.cell(row=start_row + offset, column=start_col, value=_safe_cell(label))
        ws.cell(row=start_row + offset, column=start_col + 1, value=round(hours, 2))
    ws.column_dimensions[get_column_letter(start_col)].width = 22
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 12


def _bar_chart(title: str, ws: Worksheet, start_col: int, start_row: int, count: int) -> BarChart:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Heures"
    data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + count)
    cats_ref = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + count)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    return chart


def _pie_chart(title: str, ws: Worksheet, start_col: int, start_row: int, count: int) -> PieChart:
    chart = PieChart()
    chart.title = title
    data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + count)
    cats_ref = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + count)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    return chart


def _add_bar_chart_sheet(wb: Workbook, title: str, data: list[tuple[str, float]]) -> None:
    ws = wb.create_sheet(title[:31])
    _write_summary_table(ws, 1, 1, data)
    ws.add_chart(_bar_chart(title, ws, 1, 1, len(data)), "D2")


def _add_line_chart_sheet(wb: Workbook, title: str, data: list[tuple[str, float]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(["Date", "Heures"])
    for label, hours in data:
        ws.append([_safe_cell(label), round(hours, 2)])

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Heures"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "D2")


def build_account_export(
    project: models.Project, account: models.Account, entries: list[models.TimeEntry]
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Entrées"
    _write_entries_sheet(ws, entries, include_account=False)

    total_hours = sum(entry.duration_hours for entry in entries)
    ws.append([])
    ws.append(["Total", round(total_hours, 2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    if entries:
        hours_by_category: dict[str, float] = defaultdict(float)
        hours_by_sprint: dict[str, float] = defaultdict(float)
        for entry in entries:
            hours_by_category[entry.category.name if entry.category else "Sans catégorie"] += entry.duration_hours
            hours_by_sprint[entry.sprint.name if entry.sprint else "Sans sprint"] += entry.duration_hours

        category_data = sorted(hours_by_category.items(), key=lambda item: -item[1])
        sprint_data = sorted(hours_by_sprint.items(), key=lambda item: -item[1])

        # Tableaux récapitulatifs et graphiques sur la même feuille que les entrées
        # (colonnes H et au-delà), plutôt que sur des onglets séparés — tout doit
        # être visible d'un coup d'œil.
        CATEGORY_COL, SPRINT_COL, SUMMARY_ROW = 8, 11, 1  # H, K
        _write_summary_table(ws, CATEGORY_COL, SUMMARY_ROW, category_data)
        _write_summary_table(ws, SPRINT_COL, SUMMARY_ROW, sprint_data)

        ws.add_chart(_bar_chart("Par catégorie", ws, CATEGORY_COL, SUMMARY_ROW, len(category_data)), "N2")
        ws.add_chart(
            _pie_chart("Répartition par catégorie (%)", ws, CATEGORY_COL, SUMMARY_ROW, len(category_data)), "N18"
        )
        ws.add_chart(_bar_chart("Par sprint", ws, SPRINT_COL, SUMMARY_ROW, len(sprint_data)), "N34")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_burndown_export(sprint: models.Sprint, data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Burndown"

    ws.append(["Sprint", _safe_cell(sprint.name)])
    ws.append(["Début", sprint.start_date])
    ws.append(["Fin", sprint.end_date])
    ws.append(["Total story points", data["total_points"]])
    ws.append(["US sans valorisation", data["unestimated_issue_count"]])
    ws.append([])
    header_row = ws.max_row + 1

    total_points = data["total_points"]
    total_days = (sprint.end_date - sprint.start_date).days or 1
    actual_by_date = {point["date"]: point["remaining_points"] for point in data["actual"]}
    all_dates = sorted({point["date"] for point in data["ideal"]} | set(actual_by_date))

    ws.append(["Date", "Idéal (SP restants)", "Réel (SP restants)"])
    last_actual = total_points
    for d in all_dates:
        fraction = max(0.0, min(1.0, (d - sprint.start_date).days / total_days))
        ideal_value = round(total_points * (1 - fraction), 2)
        if d in actual_by_date:
            last_actual = actual_by_date[d]
        ws.append([d, ideal_value, round(last_actual, 2)])

    chart = LineChart()
    chart.title = f"Burndown — {sprint.name}"
    chart.y_axis.title = "Story points restants"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=header_row + len(all_dates))
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(all_dates))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E2")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_project_export(
    project: models.Project, entries: list[models.TimeEntry], contributors: list[models.Account]
) -> io.BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Synthèse"

    total_hours = sum(entry.duration_hours for entry in entries)
    contributor_count = len(contributors)
    average = round(total_hours / contributor_count, 2) if contributor_count else 0.0

    ws_summary.append(["Indicateur", "Valeur"])
    ws_summary.append(["Projet", _safe_cell(project.name)])
    ws_summary.append(["Total heures", round(total_hours, 2)])
    ws_summary.append(["Nombre de contributeurs", contributor_count])
    ws_summary.append(["Nombre d'entrées", len(entries)])
    ws_summary.append(["Moyenne heures / contributeur", average])
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 30

    ws_entries = wb.create_sheet("Entrées")
    _write_entries_sheet(ws_entries, entries, include_account=True)

    hours_by_account: dict[str, float] = defaultdict(float)
    hours_by_sprint: dict[str, float] = defaultdict(float)
    hours_by_date: dict[str, float] = defaultdict(float)
    for entry in entries:
        hours_by_account[entry.account_email] += entry.duration_hours
        hours_by_sprint[entry.sprint.name if entry.sprint else "Sans sprint"] += entry.duration_hours
        hours_by_date[entry.date.isoformat()] += entry.duration_hours

    if entries:
        _add_bar_chart_sheet(wb, "Heures par membre", sorted(hours_by_account.items(), key=lambda item: -item[1]))
        _add_bar_chart_sheet(
            wb, "Répartition par sprint", sorted(hours_by_sprint.items(), key=lambda item: -item[1])
        )
        _add_line_chart_sheet(wb, "Évolution temporelle", sorted(hours_by_date.items()))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
